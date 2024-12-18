import io
import os
import shutil

import openpyxl
from django.contrib.auth import authenticate, login
from django.http import HttpResponse, HttpResponseBadRequest, FileResponse
# views.py
from django.contrib import messages

from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.core.exceptions import ValidationError
from django.shortcuts import render, get_object_or_404

from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fonts import Font
from django.contrib.auth.models import User

from djangoProject import settings
from .models import Standard, Project, TestStaff, Equipment, Sample, Regulation, Comparison, Tutorial, Message, Paper
from openpyxl import load_workbook

def homepage(request):
    return render(request, 'homepage.html')

def test_standard(request):
    # 获取所有标准
    standards = Standard.objects.all()

    # 将标准数据传递到模板
    return render(request, 'test_standard.html', {'standards': standards})
def test_search_standard(request):
    query = request.GET.get('query')
    standards = Standard.objects.filter(StandardName__icontains=query)
    return render(request, 'test_standard.html', {'standards': standards})


def message(request):
    return render(request,'message.html')

def test_project(request, standard_id):
    # 获取标准对象
    standard = get_object_or_404(Standard, id=standard_id)

    # 获取该标准下的所有项目
    projects = Project.objects.filter(standard=standard)

    # 渲染模板，传递标准和项目
    return render(request, 'test_project.html', {
        'standard': standard,
        'projects': projects,
    })

def test_search_project(request, standard_id):
    query = request.GET.get('query')
    projects = Project.objects.filter(Project__icontains=query)
    standard = get_object_or_404(Standard, id=standard_id)
    return render(request, 'test_project.html', {
        'standard': standard,
        'projects': projects,
    })

def test_detail(request,project_id):
    project = get_object_or_404(Project, id=project_id)
    equipments = Equipment.objects.filter(project_id=project_id)
    samples = Sample.objects.filter(project_id=project_id)
    tutorials = Tutorial.objects.filter(project_id=project_id)
    comparisons = Comparison.objects.filter(project_id=project_id)
    papers = Paper.objects.filter(project_id=project_id, Submitter__isnull=True)
    answers = Paper.objects.filter(project_id=project_id, Submitter=request.user.username)
    testStaffs = TestStaff.objects.filter(project_id=project_id)
    return render(request, 'test_detail.html', {'project': project,
                                                'equipments': equipments,
                                                'samples': samples,
                                                'tutorials': tutorials,
                                                'comparisons': comparisons,
                                                'papers': papers,
                                                'answers': answers,
                                                'testStaffs': testStaffs,
                                                })


def admin_message(request):
    messages = Message.objects.all()
    return render(request, 'admin_message.html', {'messages':messages})


def user_signup(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')

        # Check if passwords match
        if password != confirm_password:
            error_messages = "Passwords do not match"
            return render(request, 'user_signup.html', {'error_messages': error_messages})

        # Check if username already exists
        if User.objects.filter(username=username).exists():
            error_messages = "Username is already taken"
            return render(request, 'user_signup.html', {'error_messages': error_messages})

        # Check if email already exists
        if User.objects.filter(email=email).exists():
            error_messages = "Email is already registered"
            return render(request, 'user_signup.html', {'error_messages': error_messages})

        # Create the user
        user = User.objects.create_user(username=username, email=email, password=password)
        user.save()
        return redirect('/user_login/')

    return render(request, 'user_signup.html')


def user_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            # 如果用户存在且密码正确，则登录用户
            login(request, user)
            # 获取用户的用户名
            # username = user.username
            # request.session['username'] = username
            return render(request, 'homepage.html', {'username': username})  # 将用户名传递到模板中

        else:
            # 如果用户不存在或密码错误，则重新渲染登录页面并显示错误信息
            error_message = "  Password or Username error"
            return render(request, 'user_login.html', {'error_message': error_message})
    else:
        return render(request, 'user_login.html')

def admin_standard(request):
    # 获取所有标准
    standards = Standard.objects.all()

    # 将标准数据传递到模板
    return render(request, 'admin_standard.html', {'standards': standards})

def admin_search_standard(request):
    query = request.GET.get('query')
    standards = Standard.objects.filter(StandardName__icontains=query)
    return render(request, 'admin_standard.html', {'standards': standards})

def admin_search_project(request, standard_id):
    query = request.GET.get('query')
    projects = Project.objects.filter(Project__icontains=query)
    standard = get_object_or_404(Standard, id=standard_id)
    return render(request, 'standard_projects.html', {'projects': projects, 'standard': standard})

#下载可用于上传标准的模板.xlsx文件
def download_standard_template(request):
    # 创建一个新的工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "标准模板"

    # 在工作表中写入一些示例数据
    ws['A1'] = "broad_category"
    ws['B1'] = "category"
    ws['C1'] = "project"
    ws['D1'] = "standard_name"
    ws['E1'] = "standard_number"

    for cell in ws[1]:  # 第一行是列名
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_width = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        column_width = (max_length + 2)  # +2 for some padding
        ws.column_dimensions[cell.column_letter].width = column_width

    # 使用 io.BytesIO 对象将工作簿保存到内存中
    excel_stream = io.BytesIO()
    wb.save(excel_stream)

    # 重置流的位置到开头，以便可以读取它
    excel_stream.seek(0)

    # 设置响应头
    response = HttpResponse(
        excel_stream,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="standard_template.xlsx"'

    # 关闭流（可选，因为 HttpResponse 会在发送完数据后自动关闭它）
    # excel_stream.close()

    return response
def upload_standard(request):
    error_message = None

    # 处理文件上传
    if request.method == 'POST' and request.FILES.get('standard_file'):
        uploaded_file = request.FILES['standard_file']
        print(f"Uploaded file: {uploaded_file.name}")

        # 验证文件类型
        if not uploaded_file.name.endswith('.xlsx'):
            error_message = "只支持上传 .xlsx 格式的文件"
        else:
            try:
                # 加载Excel文件
                wb = load_workbook(uploaded_file)
                sheet = wb.active
                print("Excel file loaded successfully.")

                # 检查前几行数据
                for row in sheet.iter_rows(min_row=2, max_row=5, values_only=True):
                    print(f"Row data: {row}")  # 输出前几行数据，确保它们符合预期

                # 假设Excel中的数据从第二行开始，第一行是标题
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # 仅取前5列数据
                    row = row[:5]

                    # 过滤掉空行
                    if all(cell is None for cell in row):
                        continue  # 跳过空行

                    # 如果某个字段是空的，则跳过此行
                    if None in row:
                        print(f"Skipping row due to missing values: {row}")
                        continue

                    broad_category, category, project, standard_name, standard_number = row
                    print(f"Processing row: {broad_category}, {category}, {project}, {standard_name}, {standard_number}")

                    # 在数据库中创建新标准
                    Standard.objects.create(
                        BroadCategory=broad_category,
                        Category=category,
                        Project=project,
                        StandardName=standard_name,
                        StandardNumber=standard_number,
                    )

                print("File processed successfully.")
                return redirect('admin_standard')

            except Exception as e:
                error_message = f"文件处理出错: {e}"

    # 获取所有标准
    standards = Standard.objects.all()

    return render(request, 'admin_standard.html', {'standards': standards, 'error_message': error_message})

#下载可用于上传的项目模板文件(.xlsx)
def download_project_template(request):
    # 创建一个新的工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目模板"

    # 在工作表中写入一些示例数据
    ws['A1'] = "broad_category"
    ws['B1'] = "category"
    ws['C1'] = "project"
    ws['D1'] = "standard_name"
    ws['E1'] = "standard_number"
    ws['F1'] = "clause_number"

    for cell in ws[1]:  # 第一行是列名
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_width = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        column_width = (max_length + 2)  # +2 for some padding
        ws.column_dimensions[cell.column_letter].width = column_width

    # 使用 io.BytesIO 对象将工作簿保存到内存中
    excel_stream = io.BytesIO()
    wb.save(excel_stream)

    # 重置流的位置到开头，以便可以读取它
    excel_stream.seek(0)

    # 设置响应头
    response = HttpResponse(
        excel_stream,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="project_template.xlsx"'

    # 关闭流（可选，因为 HttpResponse 会在发送完数据后自动关闭它）
    # excel_stream.close()

    return response
def standard_projects(request, standard_id):
    # 获取标准对象
    standard = get_object_or_404(Standard, id=standard_id)

    # 获取该标准下的所有项目
    projects = Project.objects.filter(standard=standard)

    if request.method == 'POST' and request.FILES.get('project_file'):
        uploaded_file = request.FILES['project_file']

        # 验证文件类型
        if not uploaded_file.name.endswith('.xlsx'):
            error_message = "只支持上传 .xlsx 格式的文件"
        else:
            try:
                # 加载 Excel 文件
                wb = load_workbook(uploaded_file)
                sheet = wb.active
                print(f"Excel file loaded successfully. Sheet name: {sheet.title}")

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    print(f"Row data: {row}")

                    # 检查数据是否完整
                    if None in row:
                        print(f"Skipping row with empty value: {row}")
                        continue

                    broad_category, category, name, standard_name, standard_number, clause_number = row

                    # 在这里通过 standard_id 获取 Standard 实例
                    try:
                        standard = Standard.objects.get(id=standard_id)
                    except Standard.DoesNotExist:
                        print(f"Standard with id {standard_id} does not exist. Skipping project creation.")
                        continue

                    try:
                        # 在数据库中创建新项目，并将其与正确的 Standard 关联
                        project = Project.objects.create(
                            Category=category,
                            Project=name,
                            StandardName=standard_name,
                            StandardNumber=standard_number,
                            ClauseNumber=clause_number,
                            standard=standard,  # 设置外键关联
                        )
                        Tutorial.objects.create(
                            Category=category,
                            Project=name,
                            StandardName=standard_name,
                            StandardNumber=standard_number,
                            ClauseNumber=clause_number,
                            project=project  # 设置外键关联
                        )
                        Equipment.objects.create(
                            Category=category,
                            Project=name,
                            StandardName=standard_name,
                            StandardNumber=standard_number,
                            ClauseNumber=clause_number,
                            project=project  # 设置外键关联
                        )
                        Comparison.objects.create(
                            Category=category,
                            Project=name,
                            StandardName=standard_name,
                            StandardNumber=standard_number,
                            ClauseNumber=clause_number,
                            project=project  # 设置外键关联
                        )
                        Regulation.objects.create(
                            BroadCategory=broad_category,
                            Category=category,
                            Project=name,
                            StandardName=standard_name,
                            StandardNumber=standard_number,
                            ClauseNumber=clause_number,
                            project=project  # 设置外键关联
                        )
                    except Exception as e:
                        print(f"Error occurred while creating project {name}: {e}")
                        # 记录错误并向用户展示错误消息
                        messages.error(request, f"创建项目 {name} 时出错: {e}")
                        continue  # 继续处理下一个项目

                # 提示成功信息
                messages.success(request, '文件上传成功，项目数据已导入。')

            except Exception as e:
                print(f"Error occurred while processing the file: {e}")
                # 记录文件处理错误
                messages.error(request, f"文件处理出错: {e}")

    # 获取所有项目数据
    projects = Project.objects.filter(standard=standard)

    # 渲染模板，传递标准、项目
    return render(request, 'standard_projects.html', {
        'standard': standard,
        'projects': projects,
    })


def project_detail(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    equipments = Equipment.objects.filter(project_id=project_id)
    samples = Sample.objects.filter(project_id=project_id)
    tutorials = Tutorial.objects.filter(project_id=project_id)
    comparisons = Comparison.objects.filter(project_id=project_id)
    papers = Paper.objects.filter(project_id=project_id, Submitter__isnull=True)
    answers = Paper.objects.filter(project_id=project_id, Submitter__isnull=False)
    testStaffs = TestStaff.objects.filter(project_id=project_id)
    return render(request, 'project_detail.html', {'project': project,
                                                   'equipments': equipments,
                                                   'samples': samples,
                                                   'tutorials': tutorials,
                                                   'comparisons':comparisons,
                                                   'papers': papers,
                                                   'answers': answers,
                                                   'testStaffs': testStaffs,
                                                   })
def join_in_project(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    newTestStaff = TestStaff.objects.create(
        Name=request.user.username,
        Category=project.Category,
        Project=project.Project,
        StandardName=project.StandardName,
        StandardNumber=project.StandardNumber,
        ClauseNumber=project.ClauseNumber,
        project=project
    )
    newTestStaff.save()
    return redirect('test_detail', project_id)
def upload_tutorial(request, project_id):
    if request.method == 'POST' and request.FILES.get('tutorial_media'):
        name = request.POST.get('tutorial_name')
        media = request.FILES['tutorial_media']
        project = get_object_or_404(Project, id=project_id)
        print(f"name: {name}+media: {media}")
        try:
            newTutorial = Tutorial.objects.create(
                Category=project.Category,
                Project=project.Project,
                StandardName=project.StandardName,
                StandardNumber=project.StandardNumber,
                ClauseNumber=project.ClauseNumber,
                Name=name,
                Media=media,
                project_id=project_id,
            )
            newTutorial.save()
        except Exception as e:
            print(f"Error occurred while uploading tutorial {name}: {e}")
            # 记录错误并向用户展示错误消息
            messages.error(request, f"上传学习资料 {name} 时出错: {e}")
    return redirect("project_detail", project_id=project_id)
def delete_tutorial(request, project_id):
    if request.method == 'POST':
        tutorial_id = request.POST.get('tutorial_id')
        tutorial = get_object_or_404(Tutorial, id=tutorial_id)
        file_path = tutorial.Media.path
        os.remove(file_path)
        tutorial.delete()
    return redirect("project_detail", project_id=project_id)
def upload_blank_paper(request, project_id):
    name = request.POST.get('paper_name')
    file = request.FILES['paper_file']
    project = get_object_or_404(Project, id=project_id)
    try:
        newPaper = Paper.objects.create(
            Category=project.Category,
            Project=project.Project,
            StandardName=project.StandardName,
            StandardNumber=project.StandardNumber,
            ClauseNumber=project.ClauseNumber,
            Name=name,
            Paper=file,
            Type='blank',
            project_id=project_id,
        )
        newPaper.save()
    except Exception as e:
        print(f"Error occurred while adding equipment {name}: {e}")
        # 记录错误并向用户展示错误消息
        messages.error(request, f"添加设备 {name} 时出错: {e}")
    return redirect("project_detail", project_id=project_id)
def delete_paper(request, project_id):
    if request.method == 'POST':
        paper_id = request.POST.get('paper_id')
        paper = Paper.objects.get(id=paper_id)
        file_path = paper.Paper.path
        os.remove(file_path)
        paper.delete()
    return redirect("project_detail", project_id=project_id)

def add_equipment(request, project_id):
    if request.method == 'POST' and request.FILES['equipment_photo']:
        equipment = request.POST.get('equipment_name')
        manufacturer = request.POST.get('equipment_manufacturer')
        photo = request.FILES['equipment_photo']
        detail = request.POST.get('equipment_detail')
        project = get_object_or_404(Project, id=project_id)
        #print(f"{equipment}+{manufacturer}+{photo}+{detail}+{project_id}")
        try:
            newEquipment = Equipment.objects.create(
                Category=project.Category,
                Project=project.Project,
                StandardName=project.StandardName,
                StandardNumber=project.StandardNumber,
                ClauseNumber=project.ClauseNumber,
                Equipment=equipment,
                Manufacturer=manufacturer,
                Photo=photo,
                Detail=detail,
                project_id=project_id,
            )
            newEquipment.save()
        except Exception as e:
            print(f"Error occurred while adding equipment {equipment}: {e}")
            # 记录错误并向用户展示错误消息
            messages.error(request, f"添加设备 {equipment} 时出错: {e}")
    return redirect("project_detail", project_id = project_id)
def delete_equipment(request, project_id):
    equipment = Equipment.objects.get(id=request.POST.get('equipment_id'))
    file_path = equipment.Photo.path
    os.remove(file_path)
    equipment.delete()
    return redirect("project_detail", project_id = project_id)

def add_sample(request, project_id):
    if request.method == 'POST' :
        sample = request.POST.get('sample_name')
        specification = request.POST.get('sample_specification')
        manufacturer = request.POST.get('sample_manufacturer')
        batchNumber = request.POST.get('sample_batchNumber')
        try:
            newSample = Sample.objects.create(
                Sample=sample,
                Specification=specification,
                Manufacturer=manufacturer,
                BatchNumber=batchNumber,
                project_id=project_id,
            )
            newSample.save()
        except Exception as e:
            print(f"Error occurred while adding sample {sample}: {e}")
            # 记录错误并向用户展示错误消息
            messages.error(request, f"添加样品 {sample} 时出错: {e}")
    return redirect("project_detail", project_id = project_id)
def delete_sample(request, project_id):
    sample = Sample.objects.get(id=request.POST.get('sample_id'))
    sample.delete()
    return redirect("project_detail", project_id = project_id)
def download_regulation(request, project_id):
    return redirect("project_detail", project_id=project_id)
def create_comparison(request, project_id):
    if request.method == 'POST':
        name = request.POST.get('comparison_name')
        startDate = request.POST.get('comparison_startDate')
        project = get_object_or_404(Project, id=project_id)

        try:
            newComparison = Comparison.objects.create(
                Category=project.Category,
                Project=project.Project,
                StandardName=project.StandardName,
                StandardNumber=project.StandardNumber,
                ClauseNumber=project.ClauseNumber,
                Name=name,
                StartDate=startDate,
                project_id=project_id,
            )
            newComparison.save()
        except Exception as e:
            print(f"Error occurred while creating comparison {name}: {e}")
            # 记录错误并向用户展示错误消息
            messages.error(request, f"创建比对测试 {name} 时出错: {e}")
    return redirect("project_detail", project_id=project_id)
def start_comparison(request, project_id):
    if request.method == 'POST':
        comparison_id = request.POST.get('comparison_id')
        comparison = get_object_or_404(Comparison, id=comparison_id)
        comparison.State = "已开始"
        comparison.save()
    return redirect("project_detail", project_id=project_id)
def cancel_comparison(request, project_id):
    if request.method == 'POST':
        comparison_id = request.POST.get('comparison_id')
        comparison = get_object_or_404(Comparison, id=comparison_id)
        comparison.delete()
    return redirect("project_detail", project_id=project_id)

#以下为测试人员可见页面相关视图函数
def download_blank_paper(request, paper_id):
    # 获取纸张对象，若不存在则返回 404 错误
    paper = get_object_or_404(Paper, id=paper_id)

    # 确定PDF文件的路径
    file_path = paper.Paper.path  # 确保这个路径是文件的绝对路径

    # 检查文件是否存在
    if not os.path.exists(file_path):
        return HttpResponse("文件未找到", status=404)

    try:
        # 使用 FileResponse 来处理文件下载
        response = FileResponse(open(file_path, 'rb'), content_type="application/pdf")
        # 设置下载文件的名称
        response['Content-Disposition'] = f'attachment; filename="{paper.Paper.name}"'
        return response
    except Exception as e:
        # 捕捉异常并返回错误信息
        return HttpResponse(f"发生错误: {str(e)}", status=500)
def upload_paper(request, project_id):
    name = request.POST.get('paper_name')
    file = request.FILES['paper_file']
    submitter = request.POST.get('paper_submitter')
    project = get_object_or_404(Project, id=project_id)
    try:
        newPaper = Paper.objects.create(
            Category=project.Category,
            Project=project.Project,
            StandardName=project.StandardName,
            StandardNumber=project.StandardNumber,
            ClauseNumber=project.ClauseNumber,
            Submitter=submitter,
            Name=name,
            Paper=file,
            Type='blank',
            project_id=project_id,
        )
        newPaper.save()
    except Exception as e:
        print(f"Error occurred while adding equipment {name}: {e}")
        # 记录错误并向用户展示错误消息
        messages.error(request, f"添加设备 {name} 时出错: {e}")
    return redirect("test_detail", project_id=project_id)
def join_in_comparison(request, project_id):
    if request.method == 'POST':
        comparison_id = request.POST.get('comparison_id')
        comparison = get_object_or_404(Comparison, id=comparison_id)
        if comparison.Participants is None or comparison.Participants == "":
            comparison.Participants = request.user.username
        else:
            newParticipants = comparison.Participants + ',' + request.user.username
            comparison.Participants = newParticipants
        comparison.save()
    return redirect("test_detail", project_id=project_id)

